from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import math

global eloRight



biggestUpsets = {}


eloRight = 0

def eloSetter(elo_A, elo_B, winner,eloR): # takes 2 ints and a bool
    putDown_A = 0.05*(elo_A)
    putDown_B = 0.05*(elo_B)

    diff = abs(0.1*(elo_A-elo_B))

    if elo_A > elo_B:
        if winner == True: # expect scenario
            if diff >= putDown_B: # if the gap in elos is significant
                new_A = elo_A + int(0.1*(putDown_A))
                new_B = elo_B - int(0.1*(putDown_B))

            else:
                new_A = elo_A + (putDown_A - diff)
                new_B = elo_B - (putDown_B - diff) 
            
            eloR += 1
        if winner == False: # upset scenario
            new_A = elo_A - putDown_A - diff
            new_B = elo_B + putDown_B + diff
    
    
    if elo_B > elo_A:
        if winner == True: # upset scenario
            new_B = elo_B - putDown_B - diff
            new_A = elo_A + putDown_A + diff
        if winner == False: # expect scenario
            if diff >= putDown_A: # if the gap in elos is significant
                new_B = elo_B + int(0.1*(putDown_B))
                new_A = elo_A - int(0.1*(putDown_A))

            else:
                new_B = elo_B + (putDown_B - diff)
                new_A = elo_A - (putDown_A - diff)

            eloR += 1
    
    if elo_A == elo_B: # if initial elos are equal
        if winner == True:
            new_A = elo_A + putDown_A
            new_B = elo_B - putDown_B
        if winner == False:
            new_A = elo_A - putDown_A
            new_B = elo_B + putDown_B
        
        eloR += 1

    return (new_A, new_B, eloR)

def printRoundSummary(list):
    for item in list:
        print(item)

def powerRanking(list):
    topTenElos = []
    dup_list = list[:]
    topTenTeams = []
    indexes = []
    for i in range(len(list)):
        topTenElos.append(max(list))
        list.remove(max(list))

    for elo in topTenElos:
        topTenTeams.append(getTeamFullName(dup_list.index(elo)))

    for team in topTenTeams:
        print(f'{topTenTeams.index(team) + 1}.) {team} ({math.trunc(topTenElos[topTenTeams.index(team)])})')

    


def getCode(teamName):
    print(teamName)
    for i in range(2, REGISTERED_TEAMS+2):
        if (str(teamCodesWs.cell(row=i,column=1).value) == str(teamName)):
            return int(teamCodesWs.cell(row=i,column=2).value)
    
    print("ERROR: THE TEAM NAME YOU QUERIED ISN'T IN THE CODES EXCEL SHEET")
    return
            

## Code to connect to tournament excel sheet ##

wb = Workbook()
wb = load_workbook('JR-year-masterdata.xlsx')
ws = wb['Sheet1']

# connect to codesList excel sheet

teamCodesWb = Workbook()
teamCodesWb = load_workbook('TeamCodesWB.xlsx')
teamCodesWs = teamCodesWb['Sheet1']

REGISTERED_TEAMS = 0
for row in teamCodesWs:
    if not all([cell.value == None for cell in row]):
        REGISTERED_TEAMS += 1
#REGISTERED_TEAMS+=1

# Number of rounds in current season #

rounds = 0
for row in ws:
    if not all([cell.value == None for cell in row]):
        rounds += 1
rounds -= 1

elos = [1000 for i in range(REGISTERED_TEAMS-1)]
#print(elos)
def getTeamFullName(teamCode):
    for i in range(2, REGISTERED_TEAMS+2):
        if (teamCodesWs.cell(row=i,column=2).value == teamCode):
            return str(teamCodesWs.cell(row=i,column=1).value)
#def tourneyStats(tourney):
    #roundRange = []

    #for round in range(1, rounds+1):
        #if str(ws.cell(row=round+1,column=1).value) == tourney:
            #roundRange.append(round)
    
Elos_OAT_dict = {}

def highestEloOAT():
    printList = []
    values = list(Elos_OAT_dict.values())

    for i in range(5):
        for code, elo in Elos_OAT_dict.items():  # for name, age in dictionary.iteritems():  (for Python 2.x)
            if elo == max(values):
                team_name = code
        printList.append(f'{getTeamFullName(team_name)} ({max(values)})')
        values.remove(max(values))

    for item in printList:
        item = f'{printList.index(item) + 1}.) ' + item
        print(item)


def printTeamRecord(teamName):
    teamCode = list(teamCodeDict.values()).index(teamName)

    aff_losses, aff_wins, neg_losses, neg_wins, elim = 0,0,0,0,0

    for round in range(1, rounds+1):
        
        if teamCode == int(ws.cell(row=round+1,column=3).value):
            if teamCode == int(ws.cell(row=round+1,column=5).value):
                aff_wins += 1
            else:
                aff_losses += 1
            if int(ws.cell(row=round+1,column=6).value) > 0:
                elim += 1
        if teamCode == int(ws.cell(row=round+1,column=4).value):
            if teamCode == int(ws.cell(row=round+1,column=5).value):
                neg_wins += 1
            else:
                neg_losses += 1
            if int(ws.cell(row=round+1,column=6).value) > 0:
                elim += 1

        aff_rounds = aff_losses + aff_wins
        neg_rounds = neg_losses + neg_wins

    print(f'{teamName}\n\nOverall Record : {aff_wins + neg_wins}-{aff_losses + neg_losses}\nAff Record: {aff_wins}-{aff_losses}\nNeg Record: {neg_wins}-{neg_losses}\n{elim} Elimination Rounds')
## Calculate final elos ##

round_summaries = ["[Tourney] [Round] : [AFF] vs. [NEG] --> [WINNER] [FINAL ELOS]"]

## set individual elo


for round in range(1, rounds+1):
    team_aff = getCode(ws.cell(row=round+1,column=3).value)
    team_neg = getCode(ws.cell(row=round+1,column=4).value)
    team_winner = getCode(ws.cell(row=round+1,column=5).value)
    tourney = ws.cell(row=round+1,column=1).value
    roundLabel = ws.cell(row=round+1,column=2).value

    aff_elo = elos[team_aff]
    neg_elo = elos[team_neg]
    
    if team_winner == team_aff:
        winnerPC = True
    if team_winner == team_neg:
        winnerPC = False

    new_aff_elo, new_neg_elo, eloRight = eloSetter(aff_elo, neg_elo, winnerPC, eloRight)
  
    
    elos[team_aff] = new_aff_elo
    elos[team_neg] = new_neg_elo

    aff_FULL = getTeamFullName(team_aff)
    neg_FULL = getTeamFullName(team_neg)
    winner_FULL = getTeamFullName(team_winner)

    # add new elos to the OAT dictionary #

    # AFF #
    if team_aff in Elos_OAT_dict:
        if new_aff_elo > Elos_OAT_dict[team_aff]:
            Elos_OAT_dict[team_aff] = new_aff_elo
    else:
        Elos_OAT_dict[team_aff] = new_aff_elo
    
    # NEG #
    if team_neg in Elos_OAT_dict:
        if new_neg_elo > Elos_OAT_dict[team_neg]:
            Elos_OAT_dict[team_neg] = new_neg_elo
    else:
        Elos_OAT_dict[team_neg] = new_neg_elo

    round_summaries.append(f'{tourney} {roundLabel} : {aff_FULL} ({aff_elo}) vs. {neg_FULL} ({neg_elo}) --> {winner_FULL} [AFF: ({new_aff_elo}) NEG: ({new_neg_elo})]')
    
#tourneyStats("Jordan")

#powerRanking(elos)
#printRoundSummary(round_summaries)
while True:
    user = input("\nranking, team, season-summary, highest oat: ")
    print(f'So far, the elo bot has correctly predicted the winner of a round {int((eloRight/rounds)*100)}% of the time!')

    if user == "ranking":
        powerRanking(elos)
    if user == "team":
        printTeamRecord(input("Enter team code: "))
    if user == "season-summary":
        printRoundSummary(round_summaries)
    if user == "exit":
        exit()
    if user == "highest oat":
        highestEloOAT()
